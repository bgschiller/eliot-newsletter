#!/usr/bin/env python3
# /// script
# requires-python = ">=3.10"
# dependencies = [
#     "openpyxl",
# ]
# ///
"""Compare MDX schedule files against the Excel spreadsheet for timing errors.

Usage:
    uv run scripts/check-schedule.py
    uv run scripts/check-schedule.py --day 2026-07-12
"""

import argparse
import re
import sys
from pathlib import Path
from typing import Any

import openpyxl

ROOT = Path(__file__).parent.parent
SCHEDULES_DIR = ROOT / "src" / "schedules"
XLSX_PATH = ROOT / "Schedule 7_10 Draft - July Eliot 2026.xlsx"

# Map day names to MDX date strings
DAY_TO_DATE = {
    "Sunday": "2026-07-12",
    "Monday": "2026-07-13",
    "Tuesday": "2026-07-14",
    "Wednesday": "2026-07-15",
    "Thursday": "2026-07-16",
    "Friday": "2026-07-17",
    "Saturday": "2026-07-18",
}

# XLSX sheet name to date string
SHEET_TO_DATE = {
    "Sunday (0712)": "2026-07-12",
    "Monday (0713)": "2026-07-13",
    "Tuesday (0714)": "2026-07-14",
    "Wednesday (0715)": "2026-07-15",
    "Thursday (0716)": "2026-07-16",
    "Friday (0717)": "2026-07-17",
    "Saturday (0718)": "2026-07-18",
}


def xlsx_time_to_str(value: Any, past_noon: bool = False) -> str:
    """Convert an Excel time value to 24h 'H:MM' string.

    Float values like 700.0 = 7:00 AM, 100.0 = 1:00 PM, 600.0 = 6:00 PM.
    Heuristic: hour 1-6 is PM. If past_noon, hours 7-11 are also PM.
    """
    import datetime

    if value is None:
        return ""
    if isinstance(value, datetime.time):
        h, m = value.hour, value.minute
    elif isinstance(value, (int, float)):
        total_minutes = int(value)
        h = total_minutes // 100
        m = total_minutes % 100
    else:
        return str(value).strip()
    # PM heuristic
    if 1 <= h <= 6:
        h += 12
    elif past_noon and 7 <= h <= 11:
        h += 12
    if m == 0:
        return str(h)
    return f"{h}:{m:02d}"


def parse_xlsx(xlsx_path: Path) -> dict[str, list[dict]]:
    """Parse the spreadsheet, returning {date: [events]} grouped by time slot.

    Each event: {name, location, start, end}
    """
    import datetime

    wb = openpyxl.load_workbook(xlsx_path)
    events_by_date: dict[str, list[dict]] = {}

    for sheet_name, date in SHEET_TO_DATE.items():
        ws = wb[sheet_name]
        events: list[dict] = []
        current_start = ""
        past_noon = False

        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
            if row[0] is not None and str(row[0]).strip():
                current_start = xlsx_time_to_str(row[0], past_noon)
                # Detect transition to afternoon: times 100-545 or 1200+
                t = row[0]
                if isinstance(t, datetime.time):
                    mins = t.hour * 60 + t.minute
                elif isinstance(t, (int, float)):
                    mins = (int(t) // 100) * 60 + (int(t) % 100)
                    # Afternoon: 100-545 or >= 1200 (except 700-1159 = AM)
                    if (100 <= int(t) <= 545) or int(t) >= 1200:
                        past_noon = True
                else:
                    mins = 0
                if mins >= 12 * 60:  # noon or later
                    past_noon = True
            if row[1] is None:
                continue
            name = str(row[1]).strip().strip("'")  # Remove surrounding single quotes
            if not name or name == " ":
                continue
            location = str(row[2]).strip().strip("'") if row[2] else ""
            end = xlsx_time_to_str(row[3], past_noon) if row[3] else ""
            events.append(
                {
                    "name": name,
                    "location": location,
                    "start": current_start,
                    "end": end,
                }
            )

        events_by_date[date] = events

    return events_by_date


def parse_mdx(mdx_path: Path) -> list[dict]:
    """Parse an MDX schedule file, returning [{name, location, start, end}]."""

    text = mdx_path.read_text()
    events: list[dict] = []

    # Find all Timeslot blocks and their contents
    # Pattern: <Timeslot start="..."> ... </Timeslot>
    timeslot_pattern = re.compile(
        r"<Timeslot\s+start=\"([^\"]+)\"[^>]*>(.*?)</Timeslot>",
        re.DOTALL,
    )
    for match in timeslot_pattern.finditer(text):
        start = match.group(1)
        body = match.group(2)

        # Find all <li> entries within this timeslot
        li_pattern = re.compile(r"<li>(.*?)</li>", re.DOTALL)
        for li_match in li_pattern.finditer(body):
            li_text = li_match.group(1).strip()

            # Extract the event name (text before <span>)
            name_match = re.match(r"^(.*?)(?:<span|\Z)", li_text, re.DOTALL)
            name = name_match.group(1).strip() if name_match else li_text
            # Remove trailing non-breaking space entities
            name = re.sub(r"\s*\{\" \"\}\s*$", "", name)
            # Handle JSX {" "} (non-breaking space expression) in MDX
            name = re.sub(r'\s*\{" "\}\s*$', "", name)
            name = name.strip()

            # Extract location and end time from <span class="loc">
            span_match = re.search(
                r'<span\s+class="loc">(.*?)</span>', li_text, re.DOTALL
            )
            location = ""
            end = ""
            if span_match:
                span_text = span_match.group(1).strip()
                end = parse_mdx_end_time(span_text)
                # Location is everything before ", to ..." if end was found
                if end:
                    to_match = re.match(r"^(.*),\s*to\s+", span_text)
                    location = to_match.group(1).strip() if to_match else ""
                else:
                    location = span_text

            events.append(
                {
                    "name": name,
                    "location": location,
                    "start": start,
                    "end": end,
                }
            )

    return events


def normalize_name(name: str) -> str:
    """Normalize event names for fuzzy comparison."""
    n = name.strip()
    # Decode HTML entities
    n = n.replace("&eacute;", "é").replace("&amp;", "&")
    # Collapse whitespace first
    n = re.sub(r"\s+", " ", n)
    # Remove trailing punctuation
    n = n.lower().rstrip(".!")
    # Normalize common variations
    n = n.replace("morning blend café", "the morning blend")
    # Normalize all dashes to spaces
    n = re.sub(r"\s*[—–\-]\s*", " ", n)
    # Remove connectors
    n = re.sub(r"\band\b", " ", n)
    n = n.replace("&", " ")
    n = n.replace(",", " ")
    n = n.replace(":", " ")
    n = n.replace("/", " ")
    n = n.replace("(", " ")
    n = n.replace(")", " ")
    # Collapse whitespace again after removals
    n = re.sub(r"\s+", " ", n).strip()
    return n


def normalize_time(t: str) -> str:
    """Normalize time strings for comparison: '2' -> '2:00', 'noon' -> '12:00'.

    Uses 24h output: bare numbers 1-6 become PM (13:00-18:00).
    """
    t = t.strip()
    if t.lower() == "noon":
        return "12:00"
    if t.lower() == "midnight":
        return "0:00"
    # Bare number: apply same heuristic as xlsx (1-6 = PM)
    m = re.match(r"^(\d+)$", t)
    if m:
        h = int(m.group(1))
        # 7-11 = AM, 1-6 = PM, 12 = noon
        if 1 <= h <= 6:
            h += 12
        return f"{h}:00"
    # Already has colon: use as-is (assumes already in reasonable format)
    m = re.match(r"^(\d{1,2}):(\d{2})$", t)
    if m:
        h = int(m.group(1))
        if 1 <= h <= 6:
            h += 12
        return f"{h}:{m.group(2)}"
    return t


def parse_mdx_end_time(span_text: str) -> str:
    """Extract end time from MDX span text: 'Location, to HH:MM' or 'to noon'."""
    span_text = span_text.strip()
    # Match: optional location + ", to" or just "to" + time
    to_match = re.match(
        r"^(?:.*,\s*)?to\s+(\d{1,2}(?::\d{2})?|noon|midnight)\s*$",
        span_text,
        re.IGNORECASE,
    )
    if to_match:
        return to_match.group(1).strip()
    return ""


def time_to_minutes(t: str) -> int | None:
    """Convert a time string to minutes since midnight, or None if unparseable."""
    t = normalize_time(t)
    m = re.match(r"^(\d{1,2}):(\d{2})$", t)
    if m:
        return int(m.group(1)) * 60 + int(m.group(2))
    return None


# Intentional differences between xlsx and MDX (not errors)
KNOWN_DIFFERENCES: set[tuple[str, str]] = {
    # Morning Blend merged into 7:00 timeslot (was 7:30 in xlsx)
    ("the morning blend", "start"),
}

# Events where MDX intentionally omits the end time (consistent design choice)
NO_END_TIME_OK: set[str] = {
    "breakfast", "lunch", "lunch bbq", "pick up children",
}

# Events intentionally not in MDX (after-hours or post-schedule)
SKIP_FROM_MDX: set[str] = {
    "parade to salmon bake beach",
    "everyone off campus",
}


def find_best_match(
    md: dict, xlsx_list: list[dict], used_indices: set[int]
) -> tuple[dict | None, int | None]:
    """Find the best xlsx match for an MDX event by name + time proximity."""
    md_name = normalize_name(md["name"])
    md_mins = time_to_minutes(md["start"]) or 0

    best: tuple[dict | None, int | None, int] = (None, None, 9999)

    for i, xl in enumerate(xlsx_list):
        if i in used_indices:
            continue
        xl_name = normalize_name(xl["name"])

        # Must be a name match (exact or one contains the other)
        if md_name != xl_name and md_name not in xl_name and xl_name not in md_name:
            continue

        # Prefer closest start time
        xl_mins = time_to_minutes(xl["start"]) or 0
        dist = abs(md_mins - xl_mins)
        if dist < best[2]:
            best = (xl, i, dist)

    return (best[0], best[1])


def compare_day(date: str, xlsx_events: list[dict], mdx_events: list[dict]) -> list[str]:
    """Compare events for a single day. Returns list of error messages."""

    errors: list[str] = []

    # Filter xlsx events to daytime only (before 6pm, skip TBD)
    DAYTIME_CUTOFF = 18 * 60  # 6:00pm
    xlsx_relevant = []
    for e in xlsx_events:
        if e["start"].upper() == "TBD":
            continue
        start_mins = time_to_minutes(e["start"])
        if start_mins is None:
            continue
        if start_mins >= DAYTIME_CUTOFF:
            continue
        xlsx_relevant.append(e)

    used_xlsx: set[int] = set()
    matched_mdx: set[int] = set()

    # Match each MDX event to best xlsx event
    for i, md in enumerate(mdx_events):
        xl, xl_idx = find_best_match(md, xlsx_relevant, used_xlsx)

        if xl is None:
            errors.append(f"NOT IN XLSX: \"{md['name']}\" (start={md['start']})")
            continue

        used_xlsx.add(xl_idx)
        matched_mdx.add(i)

        md_name_key = normalize_name(md["name"])

        # Check start time (skip known differences)
        if (md_name_key, "start") not in KNOWN_DIFFERENCES:
            md_start = normalize_time(md["start"])
            xl_start = normalize_time(xl["start"])
            if md_start != xl_start:
                errors.append(
                    f"START TIME: \"{md['name']}\" — "
                    f"xlsx={xl['start']}, mdx={md['start']}"
                )

        # Check end time
        md_end = normalize_time(md["end"])
        xl_end = normalize_time(xl["end"])
        if xl_end:
            if md_end:
                if md_end != xl_end:
                    errors.append(
                        f"END TIME: \"{md['name']}\" — "
                        f"xlsx={xl['end']}, mdx={md['end']}"
                    )
            else:
                if normalize_name(md["name"]) not in NO_END_TIME_OK:
                    errors.append(
                        f"MISSING END TIME: \"{md['name']}\" — "
                        f"xlsx ends at {xl['end']}"
                    )

    # Report xlsx events not found in MDX
    for i, xl in enumerate(xlsx_relevant):
        if i not in used_xlsx:
            if normalize_name(xl["name"]) not in SKIP_FROM_MDX:
                errors.append(
                    f"MISSING FROM MDX: \"{xl['name']}\" (xlsx start={xl['start']})"
                )

    return errors


def main():
    parser = argparse.ArgumentParser(description="Check MDX schedule timings against xlsx")
    parser.add_argument("--day", help="Check only a specific date (e.g., 2026-07-12)")
    args = parser.parse_args()

    if not XLSX_PATH.exists():
        print(f"ERROR: xlsx not found at {XLSX_PATH}")
        sys.exit(1)

    print(f"Reading {XLSX_PATH.name}...")
    xlsx_data = parse_xlsx(XLSX_PATH)

    if args.day:
        dates = [args.day]
    else:
        dates = sorted(DAY_TO_DATE.values())

    total_errors = 0
    for date in dates:
        mdx_path = SCHEDULES_DIR / f"{date}-early.mdx"
        if not mdx_path.exists():
            print(f"\n{'=' * 60}")
            print(f"  {date} — SKIP (no MDX file)")
            continue

        print(f"\n{'=' * 60}")
        print(f"  {date} ({mdx_path.name})")

        xlsx_events = xlsx_data.get(date, [])
        if not xlsx_events:
            print("  No xlsx data for this date")
            continue

        mdx_events = parse_mdx(mdx_path)
        errors = compare_day(date, xlsx_events, mdx_events)

        if errors:
            total_errors += len(errors)
            for err in errors:
                print(f"  ❌ {err}")
        else:
            print("  ✅ All good!")

    print(f"\n{'=' * 60}")
    print(f"Total errors: {total_errors}")
    sys.exit(1 if total_errors > 0 else 0)


if __name__ == "__main__":
    main()
